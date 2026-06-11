from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
import sqlite3
import os
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'log-kios-secret-key-2024')
ADMIN_PIN = os.environ.get('ADMIN_PIN', '1234')
DB_PATH = 'visitor.db'

# ───────────────────────────────
# DB 초기화
# ───────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    # 출입자 기록 테이블
    c.execute('''
        CREATE TABLE IF NOT EXISTS visitors (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            visit_date  TEXT NOT NULL,
            time_in     TEXT NOT NULL,
            time_out    TEXT NOT NULL DEFAULT '18:00',
            affiliation TEXT NOT NULL,
            department  TEXT NOT NULL DEFAULT '',
            name        TEXT NOT NULL,
            extra_count INTEGER NOT NULL DEFAULT 0,
            purpose     TEXT NOT NULL,
            note        TEXT,
            created_at  TEXT NOT NULL
        )
    ''')
    # 사전 등록 소속 테이블
    c.execute('''
        CREATE TABLE IF NOT EXISTS affiliations (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    ''')
    # 사전 등록 부서/연구실 테이블 (소속 하위)
    c.execute('''
        CREATE TABLE IF NOT EXISTS departments (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            affiliation_id INTEGER NOT NULL,
            name           TEXT NOT NULL,
            FOREIGN KEY (affiliation_id) REFERENCES affiliations(id) ON DELETE CASCADE
        )
    ''')
    # 사전 등록 인원 테이블 (부서 하위)
    c.execute('''
        CREATE TABLE IF NOT EXISTS members (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            name          TEXT NOT NULL,
            FOREIGN KEY (department_id) REFERENCES departments(id) ON DELETE CASCADE
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# ───────────────────────────────
# 키오스크 메인
# ───────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

# ───────────────────────────────
# 단계별 등록 플로우
# ───────────────────────────────
@app.route('/step/affiliation')
def step_affiliation():
    conn = get_db()
    affiliations = conn.execute('SELECT * FROM affiliations ORDER BY name').fetchall()
    conn.close()
    return render_template('step_affiliation.html', affiliations=affiliations)

@app.route('/step/department/<int:aff_id>')
def step_department(aff_id):
    conn = get_db()
    aff = conn.execute('SELECT * FROM affiliations WHERE id=?', (aff_id,)).fetchone()
    departments = conn.execute('SELECT * FROM departments WHERE affiliation_id=? ORDER BY name', (aff_id,)).fetchall()
    conn.close()
    if not aff:
        return redirect(url_for('step_affiliation'))
    return render_template('step_department.html', aff=aff, departments=departments)

@app.route('/step/member/<int:dept_id>')
def step_member(dept_id):
    conn = get_db()
    dept = conn.execute('SELECT d.*, a.name as aff_name FROM departments d JOIN affiliations a ON d.affiliation_id=a.id WHERE d.id=?', (dept_id,)).fetchone()
    members = conn.execute('SELECT * FROM members WHERE department_id=? ORDER BY name', (dept_id,)).fetchall()
    conn.close()
    if not dept:
        return redirect(url_for('step_affiliation'))
    return render_template('step_member.html', dept=dept, members=members)

@app.route('/step/form')
def step_form():
    """최종 입력 폼 - 쿼리스트링으로 선택 값 전달"""
    aff_id  = request.args.get('aff_id', '', type=str)
    dept_id = request.args.get('dept_id', '', type=str)
    member_id = request.args.get('member_id', '', type=str)
    custom_aff  = request.args.get('custom_aff', '')
    custom_dept = request.args.get('custom_dept', '')
    custom_name = request.args.get('custom_name', '')

    conn = get_db()
    affiliation = ''
    department  = ''
    name        = ''

    if aff_id:
        row = conn.execute('SELECT name FROM affiliations WHERE id=?', (aff_id,)).fetchone()
        if row: affiliation = row['name']
    if dept_id:
        row = conn.execute('SELECT name FROM departments WHERE id=?', (dept_id,)).fetchone()
        if row: department = row['name']
    if member_id:
        row = conn.execute('SELECT name FROM members WHERE id=?', (member_id,)).fetchone()
        if row: name = row['name']
    conn.close()

    # 직접입력 값 우선
    if custom_aff:  affiliation = custom_aff
    if custom_dept: department  = custom_dept
    if custom_name: name        = custom_name

    today   = date.today().strftime('%Y-%m-%d')
    now_time = datetime.now().strftime('%H:%M')
    return render_template('step_form.html',
        affiliation=affiliation, department=department, name=name,
        today=today, now_time=now_time)

@app.route('/register', methods=['POST'])
def register():
    visit_date   = request.form.get('visit_date', '').strip()
    time_in      = request.form.get('time_in', '').strip()
    time_out     = request.form.get('time_out', '18:00').strip()
    affiliation  = request.form.get('affiliation', '').strip()
    department   = request.form.get('department', '').strip()
    name         = request.form.get('name', '').strip()
    extra_count  = request.form.get('extra_count', '0').strip()
    purpose      = request.form.get('purpose', '').strip()
    note         = request.form.get('note', '').strip()

    if not all([visit_date, time_in, affiliation, name, purpose]):
        flash('필수 항목을 모두 입력해주세요.', 'error')
        return redirect(url_for('step_affiliation'))

    try:
        extra_count = int(extra_count)
    except:
        extra_count = 0

    conn = get_db()
    conn.execute(
        'INSERT INTO visitors (visit_date,time_in,time_out,affiliation,department,name,extra_count,purpose,note,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
        (visit_date, time_in, time_out, affiliation, department, name, extra_count, purpose, note,
         datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    )
    conn.commit()
    conn.close()
    return redirect(url_for('success', name=name))

@app.route('/success')
def success():
    name = request.args.get('name', '')
    return render_template('success.html', name=name)

# ───────────────────────────────
# API (단계별 데이터)
# ───────────────────────────────
@app.route('/api/affiliations')
def api_affiliations():
    conn = get_db()
    rows = conn.execute('SELECT * FROM affiliations ORDER BY name').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/departments/<int:aff_id>')
def api_departments(aff_id):
    conn = get_db()
    rows = conn.execute('SELECT * FROM departments WHERE affiliation_id=? ORDER BY name', (aff_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/members/<int:dept_id>')
def api_members(dept_id):
    conn = get_db()
    rows = conn.execute('SELECT * FROM members WHERE department_id=? ORDER BY name', (dept_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ───────────────────────────────
# 관리자
# ───────────────────────────────
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('pin') == ADMIN_PIN:
            session['admin'] = True
            return redirect(url_for('admin'))
        flash('PIN이 올바르지 않습니다.', 'error')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('index'))

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

@app.route('/admin')
@admin_required
def admin():
    conn = get_db()
    page      = request.args.get('page', 1, type=int)
    per_page  = 20
    search    = request.args.get('search', '').strip()
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    query  = 'SELECT * FROM visitors WHERE 1=1'
    params = []
    if search:
        query  += ' AND (name LIKE ? OR affiliation LIKE ? OR purpose LIKE ? OR department LIKE ?)'
        params += [f'%{search}%']*4
    if date_from:
        query  += ' AND visit_date >= ?'
        params.append(date_from)
    if date_to:
        query  += ' AND visit_date <= ?'
        params.append(date_to)

    total = conn.execute('SELECT COUNT(*) FROM (' + query + ')', params).fetchone()[0]
    query += ' ORDER BY visit_date DESC, time_in DESC LIMIT ? OFFSET ?'
    params += [per_page, (page-1)*per_page]
    visitors = conn.execute(query, params).fetchall()
    conn.close()

    total_pages = (total + per_page - 1) // per_page
    return render_template('admin.html',
        visitors=visitors, page=page, total_pages=total_pages,
        total=total, search=search, date_from=date_from, date_to=date_to)

# 사전등록 관리
@app.route('/admin/members')
@admin_required
def admin_members():
    conn = get_db()
    affiliations = conn.execute('SELECT * FROM affiliations ORDER BY name').fetchall()
    aff_id = request.args.get('aff_id', type=int)
    dept_id = request.args.get('dept_id', type=int)
    departments, members = [], []
    sel_aff, sel_dept = None, None
    if aff_id:
        departments = conn.execute('SELECT * FROM departments WHERE affiliation_id=? ORDER BY name', (aff_id,)).fetchall()
        sel_aff = conn.execute('SELECT * FROM affiliations WHERE id=?', (aff_id,)).fetchone()
    if dept_id:
        members = conn.execute('SELECT * FROM members WHERE department_id=? ORDER BY name', (dept_id,)).fetchall()
        sel_dept = conn.execute('SELECT * FROM departments WHERE id=?', (dept_id,)).fetchone()
        if not aff_id and sel_dept:
            aff_id = sel_dept['affiliation_id']
            departments = conn.execute('SELECT * FROM departments WHERE affiliation_id=? ORDER BY name', (aff_id,)).fetchall()
            sel_aff = conn.execute('SELECT * FROM affiliations WHERE id=?', (aff_id,)).fetchone()
    conn.close()
    return render_template('admin_members.html',
        affiliations=affiliations, departments=departments, members=members,
        sel_aff=sel_aff, sel_dept=sel_dept)

@app.route('/admin/members/add_aff', methods=['POST'])
@admin_required
def add_affiliation():
    name = request.form.get('name','').strip()
    if name:
        conn = get_db()
        try:
            conn.execute('INSERT INTO affiliations (name) VALUES (?)', (name,))
            conn.commit()
        except: pass
        conn.close()
    return redirect(url_for('admin_members'))

@app.route('/admin/members/add_dept', methods=['POST'])
@admin_required
def add_department():
    aff_id = request.form.get('aff_id', type=int)
    name   = request.form.get('name','').strip()
    if aff_id and name:
        conn = get_db()
        conn.execute('INSERT INTO departments (affiliation_id, name) VALUES (?,?)', (aff_id, name))
        conn.commit()
        conn.close()
    return redirect(url_for('admin_members', aff_id=aff_id))

@app.route('/admin/members/add_member', methods=['POST'])
@admin_required
def add_member():
    dept_id = request.form.get('dept_id', type=int)
    name    = request.form.get('name','').strip()
    aff_id  = request.form.get('aff_id', type=int)
    if dept_id and name:
        conn = get_db()
        conn.execute('INSERT INTO members (department_id, name) VALUES (?,?)', (dept_id, name))
        conn.commit()
        conn.close()
    return redirect(url_for('admin_members', aff_id=aff_id, dept_id=dept_id))

@app.route('/admin/members/del_aff/<int:aid>', methods=['POST'])
@admin_required
def del_affiliation(aid):
    conn = get_db()
    conn.execute('PRAGMA foreign_keys=ON')
    conn.execute('DELETE FROM affiliations WHERE id=?', (aid,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_members'))

@app.route('/admin/members/del_dept/<int:did>', methods=['POST'])
@admin_required
def del_department(did):
    conn = get_db()
    dept = conn.execute('SELECT affiliation_id FROM departments WHERE id=?', (did,)).fetchone()
    aff_id = dept['affiliation_id'] if dept else None
    conn.execute('PRAGMA foreign_keys=ON')
    conn.execute('DELETE FROM departments WHERE id=?', (did,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_members', aff_id=aff_id))

@app.route('/admin/members/del_member/<int:mid>', methods=['POST'])
@admin_required
def del_member(mid):
    conn = get_db()
    m = conn.execute('SELECT m.department_id, d.affiliation_id FROM members m JOIN departments d ON m.department_id=d.id WHERE m.id=?', (mid,)).fetchone()
    dept_id = m['department_id'] if m else None
    aff_id  = m['affiliation_id'] if m else None
    conn.execute('DELETE FROM members WHERE id=?', (mid,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_members', aff_id=aff_id, dept_id=dept_id))

@app.route('/admin/delete/<int:vid>', methods=['POST'])
@admin_required
def delete_visitor(vid):
    conn = get_db()
    conn.execute('DELETE FROM visitors WHERE id=?', (vid,))
    conn.commit()
    conn.close()
    flash('기록이 삭제되었습니다.')
    return redirect(url_for('admin'))

@app.route('/admin/delete_all', methods=['POST'])
@admin_required
def delete_all():
    conn = get_db()
    conn.execute('DELETE FROM visitors')
    conn.commit()
    conn.close()
    flash('모든 기록이 삭제되었습니다.')
    return redirect(url_for('admin'))

@app.route('/admin/export')
@admin_required
def export():
    conn = get_db()
    rows = conn.execute('SELECT * FROM visitors ORDER BY visit_date DESC, time_in DESC').fetchall()
    conn.close()

    os.makedirs('export_logs', exist_ok=True)
    fname = f'export_logs/visitors_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '출입자 기록'

    headers = ['No','날짜','입실','퇴실','소속','부서/연구실','이름','동반인원','방문목적','기타','등록시각']
    hfill  = PatternFill('solid', fgColor='0D2B6B')
    hfont  = Font(bold=True, color='FFFFFF')
    thin   = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, row in enumerate(rows, 1):
        data = [i, row['visit_date'], row['time_in'], row['time_out'],
                row['affiliation'], row['department'], row['name'],
                f"+{row['extra_count']}명" if row['extra_count'] else '1명',
                row['purpose'], row['note'] or '', row['created_at']]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col<=4 else 'left')

    for col, w in enumerate([6,12,8,8,14,14,10,8,14,16,18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(fname)
    resp = send_file(os.path.abspath(fname), as_attachment=True,
                     download_name=os.path.basename(fname))
    @resp.call_on_close
    def cleanup():
        try: os.remove(fname)
        except: pass
    return resp

@app.route('/offline')
def offline():
    return render_template('offline.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
